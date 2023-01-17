"""
проверя скачанные парсером снимки нового проспекта за месяц и генерю
файл отчета с превью
"""

from openpyxl import Workbook
from openpyxl.drawing.image import Image
import os
from datetime import datetime
from folder_in_MY_documents import make_documets_folder

filename_data = datetime.now().day
month_name = datetime.now().month
current_year = datetime.now().year

# way_to_files = f"/Volumes/big4photo/Documents/NewProspect/{current_year}_{month_name}"  # путь к папке с изображениями
way_to_files = f"{make_documets_folder('NewProspect')}/{current_year}_{month_name}"  # путь к папке с изображениями
names = os.listdir(way_to_files)

wb = Workbook()
ws = wb.active
ws.title = "Sheet with image"  # задаю название вкладки
ws.column_dimensions['C'].width = 50  # задаю шрину колонки
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 100
ws.column_dimensions['D'].width = 20

count = 0
for name in names:
    if name.endswith('JPG') or name.endswith('jpg'):
        count += 1
        ws.row_dimensions[count].height = 150  # задаю высоту столбца
        print(f"{way_to_files}/{name}")
        img = Image(f"{way_to_files}/{name}")
        resize_height = img.height // 3  # уменьшая рарешение в два раза
        resize_width = img.width // 3  # уменьшая рарешение в два раза

        img.width = resize_width  # устанавливаю размер превью
        img.height = resize_height  # устанавливаю размер превью
        ws.add_image(img, f'C{count}')

        ws[f'A{count}'] = name.split("__")[0]
        ws[f'B{count}'] = name.split("__")[1][:-4]
        ws[f'D{count}'] = 500

wb.save(f"{way_to_files}/report_{filename_data}.xlsx")
